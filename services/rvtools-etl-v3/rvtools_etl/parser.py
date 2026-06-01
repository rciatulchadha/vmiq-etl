"""
RVTools Excel parser v3 — exact column names confirmed from file.

Column naming convention: v{Sheet}{Field}
  vInfo:      vInfoVMName, vInfoPowerstate, vInfoCPUs, vInfoMemory ...
  vHost:      vHostName, vHostCluster, vHostCpuModel, vHostMemorySize ...
  vCluster:   vClusterName, vClusterNumHosts, vClusterTotalCpu ...
  vDatastore: vDatastoreName, vDatastoreCapacity, vDatastoreFreeSpace ...
  vDisk:      vDiskVMName, vDiskRaw, vDiskMode, vDiskCapacityMiB ...
  vSnapshot:  vSnapshotVMName ...

Important notes:
  - vNIC sheet contains HOST NIC info (vNicHostName), NOT VM NIC info
  - VM network info is in vNetwork sheet (used for dvSwitch detection)
  - vDatastore capacity/freespace are in MB
  - vInfoProvisioned (disk) is in MB
  - vHostMemorySize is in bytes (divide by 1024/1024 for MB)
  - vCluster has no VM count column
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

log = logging.getLogger("rvtools-etl.parser")


def _v(row: dict, *keys, default=None) -> Any:
    """Try multiple column name variants, return first non-empty match."""
    for key in keys:
        val = row.get(key)
        if val is not None and str(val).strip() not in ("", "None"):
            return val
    return default


def _int(v, default=0) -> int:
    try:
        return int(float(str(v))) if v is not None else default
    except (ValueError, TypeError):
        return default


def _float(v, default=0.0) -> float:
    try:
        return float(str(v)) if v is not None else default
    except (ValueError, TypeError):
        return default


def _str(v, default=None) -> str:
    if v is None:
        return default
    s = str(v).strip()
    return s if s else default


def _bool(v) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    return str(v).strip().lower() in ("true", "yes", "1")


def _sheet_to_dicts(ws) -> list[dict]:
    """Row 1 = headers, Row 2+ = data."""
    headers = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [
                str(h).strip() if h is not None else f"col_{j}"
                for j, h in enumerate(row)
            ]
        else:
            if all(v is None for v in row):
                continue
            rows.append(dict(zip(headers, row)))
    return rows


class RVToolsParser:
    def _init_(self, filepath: Path, sheet_config: dict,
                 estate: str = "DEFAULT"):
        self.filepath     = filepath
        self.sheet_config = sheet_config
        self.estate       = estate
        self._wb          = None

    def _load(self):
        log.info(f"Loading workbook: {self.filepath.name}")
        self._wb = openpyxl.load_workbook(
            self.filepath, read_only=True, data_only=True)
        log.info(f"Sheets found: {self._wb.sheetnames}")

    def _sheet(self, name: str) -> list[dict]:
        if name not in self._wb.sheetnames:
            log.warning(f"Sheet '{name}' not found — skipping.")
            return []
        ws = self._wb[name]
        rows = _sheet_to_dicts(ws)
        log.info(f"  {name}: {len(rows)} rows")
        return rows

    def parse(self) -> dict:
        self._load()

        raw_vms        = self._sheet(self.sheet_config["vms"])
        raw_hosts      = self._sheet(self.sheet_config["hosts"])
        raw_clusters   = self._sheet(self.sheet_config["clusters"])
        raw_datastores = self._sheet(self.sheet_config["datastores"])
        raw_disks      = self._sheet(self.sheet_config["disks"])
        raw_snapshots  = self._sheet(self.sheet_config["snapshots"])

        # vNIC sheet has HOST nics, not VM nics — use vNetwork for VM info
        raw_network = self._sheet("vNetwork") \
            if "vNetwork" in self._wb.sheetnames else []

        self._wb.close()

        if not raw_vms:
            raise ValueError(
                f"Sheet '{self.sheet_config['vms']}' has no data rows. "
                f"Check file has VM data."
            )

        snap_vms  = self._snapshot_vms(raw_snapshots)
        disk_info = self._disk_info(raw_disks)
        nic_info  = self._nic_info(raw_network)

        # Derive vCenter name from filename if column missing
        filename_vc = self.filepath.stem.replace("_latest", "")

        vcenters   = self._vcenters(raw_vms, fallback=filename_vc)
        clusters   = self._clusters(raw_clusters)
        hosts      = self._hosts(raw_hosts)
        vms        = self._vms(raw_vms, snap_vms, disk_info, nic_info)
        datastores = self._datastores(raw_datastores)

        return {
            "estate":     self.estate,
            "vcenters":   vcenters,
            "clusters":   clusters,
            "hosts":      hosts,
            "vms":        vms,
            "datastores": datastores,
        }

    # ── vCenters ───────────────────────────────────────────────

    def _vcenters(self, rows: list, fallback: str) -> list[dict]:
        names = set()
        for row in rows:
            vc = _str(_v(row, "vInfoVISDKServer"))
            if vc:
                names.add(vc)
        if not names:
            log.warning(f"No vInfoVISDKServer column found — using filename: '{fallback}'")
            names.add(fallback)
        result = [{"name": n, "fqdn": n, "estate": self.estate}
                  for n in sorted(names)]
        log.info(f"vCenters detected: {[v['name'] for v in result]}")
        return result

    # ── Snapshots ──────────────────────────────────────────────

    def _snapshot_vms(self, rows: list) -> set:
        """Return set of VM names that have at least one snapshot."""
        return {
            _str(_v(row, "vSnapshotVMName")).lower()
            for row in rows
            if _v(row, "vSnapshotVMName")
        }

    # ── Disk info ──────────────────────────────────────────────

    def _disk_info(self, rows: list) -> dict:
        """
        Per-VM disk flags from vDisk sheet.
        vDiskRaw = True/False (RDM disk)
        vDiskMode = string containing 'independent' for independent disks
        """
        info = {}
        for row in rows:
            vm = _str(_v(row, "vDiskVMName"))
            if not vm:
                continue
            key = vm.lower()
            if key not in info:
                info[key] = {"has_rdm": False, "has_independent": False}

            raw  = _v(row, "vDiskRaw")
            mode = _str(_v(row, "vDiskMode"), "")

            # vDiskRaw is boolean True/False in this RVTools version
            if _bool(raw):
                info[key]["has_rdm"] = True
            if "independent" in mode.lower():
                info[key]["has_independent"] = True

        return info

    # ── NIC info ───────────────────────────────────────────────

    def _nic_info(self, rows: list) -> dict:
        """
        VM network info from vNetwork sheet (not vNIC which is host NICs).
        Detects dvSwitch usage per VM.
        """
        info = {}
        for row in rows:
            vm = _str(_v(row, "vNetworkVMName", "VM", "VM Name"))
            if not vm:
                continue
            key = vm.lower()
            if key not in info:
                adapter = _str(_v(row,
                    "vNetworkAdapterType", "vNetworkType",
                    "Type", "Adapter Type"), "")
                info[key] = {"net_adapter": adapter, "is_dvswitch": False}
            network = _str(_v(row,
                "vNetworkNetwork", "vNetworkPortgroup",
                "Network", "Port Group"), "")
            if "dvportgroup" in network.lower() or "dvs" in network.lower():
                info[key]["is_dvswitch"] = True
        return info

    # ── Clusters ───────────────────────────────────────────────

    def _clusters(self, rows: list) -> list[dict]:
        clusters = []
        seen = set()
        for row in rows:
            name = _str(_v(row, "vClusterName"))
            if not name or name in seen:
                continue
            seen.add(name)

            # vClusterTotalMemory is in bytes — convert to MB
            mem_bytes = _float(_v(row, "vClusterTotalMemory"))
            mem_mb = int(mem_bytes / 1024 / 1024) if mem_bytes > 1048576 \
                else int(mem_bytes)  # already in MB if small value

            clusters.append({
                "name":          name,
                "vcenter":       _str(_v(row, "vClusterVISDKServer")),
                "cpu_total_mhz": _int(_v(row, "vClusterTotalCpu")),
                "mem_total_mb":  mem_mb,
                "host_count":    _int(_v(row, "vClusterNumHosts")),
                "vm_count":      0,  # not in vCluster sheet — calculated later
            })
        return clusters

    # ── Hosts ──────────────────────────────────────────────────

    def _hosts(self, rows: list) -> list[dict]:
        hosts = []
        seen = set()
        for row in rows:
            name = _str(_v(row, "vHostName"))
            if not name or name in seen:
                continue
            seen.add(name)

            cpu_model = _str(_v(row, "vHostCpuModel"), "")

            # vHostMemorySize is in bytes — convert to MB
            mem_bytes = _float(_v(row, "vHostMemorySize"))
            mem_mb = int(mem_bytes / 1024 / 1024) if mem_bytes > 1048576 \
                else int(mem_bytes)

            hosts.append({
                "name":             name,
                "vcenter":          _str(_v(row, "vHostVISDKServer")),
                "cluster":          _str(_v(row, "vHostCluster")),
                "esxi_version":     _str(_v(row, "vHostFullName")),
                "esxi_build":       None,  # not a separate column
                "model":            _str(_v(row, "vHostModel")),
                "vendor":           _str(_v(row, "vHostVendor")),
                "cpu_sockets":      _int(_v(row, "vHostNumCpu")),
                "cpu_cores":        _int(_v(row, "vHostNumCpuCores")),
                "cpu_threads":      _int(_v(row, "vHostCoresPerCPU")),
                "mem_total_mb":     mem_mb,
                "powerstate":       "poweredOn",  # no power column in vHost
                "connection_state": _str(_v(row, "vHostConfigStatus"), "connected"),
                "is_in_maintenance": _bool(_v(row, "vHostinMaintenanceMode")),
                "is_intel_cpu":     "intel" in cpu_model.lower() if cpu_model else True,
                "mgmt_ip":          None,  # not in vHost, in vSC_VMK sheet
            })
        return hosts

    # ── Virtual Machines ───────────────────────────────────────

    def _vms(self, rows, snap_vms, disk_info, nic_info) -> list[dict]:
        vms = []
        for row in rows:
            name = _str(_v(row, "vInfoVMName"))
            if not name:
                continue

            key   = name.lower()
            dinfo = disk_info.get(key, {})
            ninfo = nic_info.get(key, {})

            # Memory — vInfoMemory is in MB
            mem_mb = _int(_v(row, "vInfoMemory"))

            # Disk — vInfoProvisioned is in MB, convert to GB
            disk_mb = _float(_v(row, "vInfoProvisioned", "vInfoInUse"))
            disk_gb = round(disk_mb / 1024, 2) if disk_mb else 0

            # FT state
            ft = _str(_v(row, "vInfoFaultToleranceState"), "")
            is_ft = ft.lower() not in ("", "none", "notconfigured", "false", "no")

            # Created date
            created = _v(row, "vInfoCreateDate")
            if isinstance(created, str):
                for fmt in ("%m/%d/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S",
                            "%m/%d/%Y", "%Y-%m-%d"):
                    try:
                        created = datetime.strptime(created, fmt)
                        break
                    except ValueError:
                        continue
                else:
                    created = None

            powerstate = _str(_v(row, "vInfoPowerstate"), "unknown")

            # USB — check vInfoNICs or use vUSB sheet separately
            # For now derive from disk info
            has_usb = False  # populated from vUSB sheet in future

            vms.append({
                "name":            name,
                "vcenter":         _str(_v(row, "vInfoVISDKServer")),
                "cluster":         _str(_v(row, "vInfoCluster")),
                "host":            _str(_v(row, "vInfoHost")),
                "powerstate":      powerstate,
                "os_type":         _str(_v(row, "vInfoOSTools")),
                "os_fullname":     _str(_v(row, "vInfoOS")),
                "vcpus":           _int(_v(row, "vInfoCPUs")),
                "mem_mb":          mem_mb,
                "disk_total_gb":   disk_gb,
                "ip_address":      _str(_v(row, "vInfoPrimaryIPAddress")),
                "tools_status":    _str(_v(row, "vInfoToolsStatus")),
                "tools_version":   _str(_v(row, "vInfoToolsVersion")),
                "hw_version":      _str(_v(row, "VInfoVersion")),
                "has_snapshots":   key in snap_vms,
                "has_rdm":         dinfo.get("has_rdm", False),
                "has_independent": dinfo.get("has_independent", False),
                "has_usb":         has_usb,
                "has_cdrom":       False,  # from vCD sheet — future
                "is_suspended":    powerstate.lower() == "suspended",
                "is_ft":           is_ft,
                "net_adapter":     ninfo.get("net_adapter"),
                "is_dvswitch":     ninfo.get("is_dvswitch", False),
                "annotation":      _str(_v(row, "vInfoNotes")),
                "created_date":    created,
            })
        return vms

    # ── Datastores ─────────────────────────────────────────────

    def _datastores(self, rows: list) -> list[dict]:
        datastores = []
        seen = set()
        for row in rows:
            name = _str(_v(row, "vDatastoreName"))
            if not name or name in seen:
                continue
            seen.add(name)

            # vDatastoreCapacity and vDatastoreFreeSpace are in MB
            cap_mb  = _float(_v(row, "vDatastoreCapacity"))
            free_mb = _float(_v(row, "vDatastoreFreeSpace"))
            cap_gb  = round(cap_mb  / 1024, 2) if cap_mb  else 0
            free_gb = round(free_mb / 1024, 2) if free_mb else 0
            used_pct = round((1 - free_mb / cap_mb) * 100, 1) \
                if cap_mb and cap_mb > 0 else 0

            # VM count — vDatastoreVMs is comma-separated VM names
            # vDatastoreVMsTotal is the count
            vm_count = _int(_v(row, "vDatastoreVMsTotal", "vDatastoreVMs"))

            datastores.append({
                "name":        name,
                "vcenter":     _str(_v(row, "vDatastoreVISDKServer")),
                "type":        _str(_v(row, "vDatastoreType")),
                "capacity_gb": cap_gb,
                "free_gb":     free_gb,
                "used_pct":    used_pct,
                "vm_count":    vm_count,
            })
        return datastores
